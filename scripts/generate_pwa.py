#!/usr/bin/env python3
"""从 travel.xlsx 的新版旅行 Sheet 生成手机友好的离线 HTML。"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import struct
import sys
import zlib
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


DEFAULT_WORKBOOK = "travel.xlsx"
DEFAULT_SHEET = "26暑期云南"
HEADERS = [
    "日期", "开始", "结束", "类型", "项目", "路线 / 备注",
    "单价", "数量", "消费", "开售 / 预约时间", "状态", "地点 / 地图关键词",
]
TODO_STATUSES = {"待购买", "待预约"}
TYPE_NAMES = {"交通", "住宿", "游览", "餐饮", "日常", "购物", "其他"}


@dataclass
class TripItem:
    row: int
    day: date
    start: time | None
    end: time | None
    kind: str
    title: str
    note: str
    unit_price: float | None
    quantity: float | None
    cost: float | None
    booking_at: datetime | date | None
    status: str
    location: str


@dataclass
class TripDay:
    day: date
    label: str
    items: list[TripItem]


WEEKDAYS = "一二三四五六日"


def as_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    return None


def as_time(value: Any, epoch: datetime) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        if isinstance(converted, datetime):
            return converted.time().replace(microsecond=0)
        if isinstance(converted, time):
            return converted.replace(microsecond=0)
    return None


def as_booking_time(value: Any, epoch: datetime) -> datetime | date | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        if isinstance(converted, (datetime, date)):
            return converted
    return None


def as_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def calculate_cost(formula_value: Any, cached_value: Any, unit: Any, quantity: Any) -> float | None:
    cached = as_number(cached_value)
    if cached is not None:
        return cached
    direct = as_number(formula_value)
    if direct is not None:
        return direct
    unit_number = as_number(unit)
    quantity_number = as_number(quantity)
    if unit_number is not None and quantity_number is not None:
        return unit_number * quantity_number
    return None


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def default_day_label(day: date) -> str:
    return f"{day.month}月{day.day}日 周{WEEKDAYS[day.weekday()]}"


def validate_sheet(sheet: Any) -> None:
    actual = [clean_text(sheet.cell(4, column).value) for column in range(1, len(HEADERS) + 1)]
    if actual != HEADERS:
        raise ValueError(
            f"Sheet“{sheet.title}”不是新版模板：第 4 行应为 {' / '.join(HEADERS)}"
        )


def read_trip(workbook_path: Path, sheet_name: str) -> tuple[str, date | None, date | None, list[TripDay]]:
    formulas = load_workbook(workbook_path, data_only=False)
    cached = load_workbook(workbook_path, data_only=True)
    try:
        if sheet_name not in formulas.sheetnames:
            raise ValueError(f"找不到 Sheet“{sheet_name}”")
        sheet = formulas[sheet_name]
        cached_sheet = cached[sheet_name]
        validate_sheet(sheet)

        title = clean_text(sheet["A1"].value) or sheet_name
        start_day = as_date(sheet["B2"].value, formulas.epoch)
        end_day = as_date(sheet["C2"].value, formulas.epoch)
        separators: dict[date, str] = {}
        items_by_day: dict[date, list[TripItem]] = {}

        merged_rows = {
            cell_range.min_row
            for cell_range in sheet.merged_cells.ranges
            if cell_range.min_col == 1 and cell_range.max_col >= 11
        }
        for row in range(5, sheet.max_row + 1):
            if row in merged_rows:
                text = clean_text(sheet.cell(row, 1).value)
                match = re.search(r"(\d{1,2})月(\d{1,2})日", text)
                if match:
                    year = start_day.year if start_day else date.today().year
                    try:
                        separators[date(year, int(match.group(1)), int(match.group(2)))] = text
                    except ValueError:
                        pass
                continue

            day = as_date(sheet.cell(row, 1).value, formulas.epoch)
            values = [sheet.cell(row, column).value for column in range(1, len(HEADERS) + 1)]
            if day is None or not any(value not in (None, "") for value in values[1:]):
                continue

            unit = sheet.cell(row, 7).value
            quantity = sheet.cell(row, 8).value
            item = TripItem(
                row=row,
                day=day,
                start=as_time(sheet.cell(row, 2).value, formulas.epoch),
                end=as_time(sheet.cell(row, 3).value, formulas.epoch),
                kind=clean_text(sheet.cell(row, 4).value) or "其他",
                title=clean_text(sheet.cell(row, 5).value) or "未命名事项",
                note=clean_text(sheet.cell(row, 6).value),
                unit_price=as_number(unit),
                quantity=as_number(quantity),
                cost=calculate_cost(
                    sheet.cell(row, 9).value,
                    cached_sheet.cell(row, 9).value,
                    unit,
                    quantity,
                ),
                booking_at=as_booking_time(sheet.cell(row, 10).value, formulas.epoch),
                status=clean_text(sheet.cell(row, 11).value) or "无需预订",
                location=clean_text(sheet.cell(row, 12).value),
            )
            items_by_day.setdefault(day, []).append(item)

        days = [
            TripDay(day=day, label=separators.get(day, default_day_label(day)), items=items)
            for day, items in sorted(items_by_day.items())
        ]
        if not days:
            raise ValueError(f"Sheet“{sheet_name}”中没有可生成的行程事项")
        return title, start_day, end_day, days
    finally:
        formulas.close()
        cached.close()


def esc(value: Any) -> str:
    return html.escape(str(value), quote=True)


def money(value: float) -> str:
    return f"¥{value:,.2f}"


def time_range(item: TripItem) -> str:
    start = item.start.strftime("%H:%M") if item.start else "待定"
    if not item.end:
        return start
    return f"{start}–{item.end.strftime('%H:%M')}"


def booking_text(value: datetime | date) -> str:
    if isinstance(value, datetime) and value.time() != time(0, 0):
        return value.strftime("%Y-%m-%d %H:%M")
    return value.strftime("%Y-%m-%d")


def slug_class(value: str, allowed: set[str], fallback: str) -> str:
    return value if value in allowed else fallback


def render_item(item: TripItem) -> str:
    kind = slug_class(item.kind, TYPE_NAMES, "其他")
    todo = item.status in TODO_STATUSES
    details: list[str] = []
    if item.note:
        details.append(f'<div class="note">{esc(item.note)}</div>')
    if item.booking_at:
        details.append(
            f'<div class="booking">⏰ 开售 / 预约：{esc(booking_text(item.booking_at))}</div>'
        )
    if item.cost is not None:
        price_detail = ""
        if item.unit_price is not None and item.quantity is not None:
            price_detail = f'<span class="price-detail">{money(item.unit_price)} × {item.quantity:g}</span>'
        details.append(f'<div class="cost">{price_detail}<strong>{money(item.cost)}</strong></div>')

    location = ""
    actions = ""
    if item.location:
        location = f'<div class="location">📍 {esc(item.location)}</div>'
        actions = (
            f'<div class="actions"><button type="button" class="map-open" '
            f'data-location="{esc(item.location)}">打开地图</button>'
            f'<button type="button" class="copy" data-copy="{esc(item.location)}">复制地点</button></div>'
        )
    copy_text = "\n".join(
        part for part in (
            f"{item.day:%Y-%m-%d} {time_range(item)}",
            item.title,
            item.location,
            item.note,
        ) if part
    )
    return f'''<article class="item type-{esc(kind)}{' todo' if todo else ''}" data-todo="{'1' if todo else '0'}" data-start="{item.day.isoformat()}T{item.start.strftime('%H:%M') if item.start else ''}" data-end="{item.day.isoformat()}T{item.end.strftime('%H:%M') if item.end else ''}">

      <div class="time">{esc(time_range(item))}</div>
      <div class="item-body">
        <div class="item-heading"><span class="type">{esc(item.kind)}</span><h3>{esc(item.title)}</h3></div>
        {''.join(details)}{location}
        <div class="item-footer"><span class="status status-{esc(item.status)}">{esc(item.status)}</span><button type="button" class="copy-item" data-copy="{esc(copy_text)}">复制事项</button>{actions}</div>
      </div>
    </article>'''


def render_html(title: str, sheet_name: str, start_day: date | None, end_day: date | None, days: list[TripDay]) -> str:
    all_items = [item for day in days for item in day.items]
    total_cost = sum(item.cost or 0 for item in all_items)
    todo_count = sum(item.status in TODO_STATUSES for item in all_items)
    date_text = ""
    if start_day and end_day:
        date_text = f"{start_day:%Y-%m-%d} 至 {end_day:%Y-%m-%d}"

    nav = "".join(
        f'<a href="#day-{day.day.isoformat()}" data-day-link="{day.day.isoformat()}"><b>{day.day.day}</b><span>{day.day.month}月</span></a>'
        for day in days
    )
    sections: list[str] = []
    for trip_day in days:
        day_cost = sum(item.cost or 0 for item in trip_day.items)
        day_todos = sum(item.status in TODO_STATUSES for item in trip_day.items)
        sections.append(f'''<section class="day" id="day-{trip_day.day.isoformat()}" data-date="{trip_day.day.isoformat()}">
  <header class="day-header">
    <div><span class="day-date">{trip_day.day:%m月%d日}</span><h2>{esc(trip_day.label)}</h2></div>
    <div class="day-summary"><strong>{money(day_cost)}</strong><span>{len(trip_day.items)} 项{f' · {day_todos} 待办' if day_todos else ''}</span></div>
  </header>
  <div class="timeline">{''.join(render_item(item) for item in trip_day.items)}</div>
</section>''')

    return f'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<meta name="color-scheme" content="light">
<meta name="theme-color" content="#2563eb">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="apple-mobile-web-app-title" content="旅行计划">
<meta name="travel-version" content="{{VERSION}}">
<link rel="manifest" href="manifest.webmanifest">
<link rel="icon" href="icons/icon-192.png">
<link rel="apple-touch-icon" href="icons/apple-touch-icon.png">
<title>{esc(title)}</title>
<style>
:root{{--bg:#f4f7fb;--card:#fff;--ink:#172033;--muted:#667085;--line:#dbe4ef;--primary:#2563eb;--shadow:0 8px 24px rgba(30,55,90,.08)}}
*{{box-sizing:border-box}}html{{--sticky-offset:calc(82px + env(safe-area-inset-top));scroll-behavior:smooth;scroll-padding-top:var(--sticky-offset)}}body{{margin:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif;-webkit-text-size-adjust:100%;text-rendering:optimizeLegibility}}button,a{{font:inherit;-webkit-tap-highlight-color:transparent}}.hero{{padding:calc(28px + env(safe-area-inset-top)) max(18px,env(safe-area-inset-right)) 20px max(18px,env(safe-area-inset-left));background:linear-gradient(135deg,#173b73,#2563eb);color:#fff}}.hero-inner,main{{max-width:760px;margin:auto}}.eyebrow{{margin:0 0 7px;font-size:12px;opacity:.76}}h1{{font-size:25px;line-height:1.25;margin:0}}.date-range{{margin:8px 0 0;opacity:.86;font-size:14px}}.stats{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:20px}}.stat{{padding:13px 14px;border:1px solid rgba(255,255,255,.18);border-radius:14px;background:rgba(255,255,255,.12);backdrop-filter:blur(6px)}}.stat span{{display:block;font-size:12px;opacity:.8}}.stat strong{{display:block;margin-top:3px;font-size:20px}}.toolbar{{position:sticky;z-index:20;top:0;background:rgba(244,247,251,.94);border-bottom:1px solid var(--line);-webkit-backdrop-filter:blur(12px);backdrop-filter:blur(12px)}}.toolbar-inner{{display:flex;align-items:center;gap:10px;max-width:760px;margin:auto;padding:9px max(14px,env(safe-area-inset-right)) 9px max(14px,env(safe-area-inset-left))}}.day-nav{{display:flex;gap:7px;overflow-x:auto;scrollbar-width:none;flex:1}}.day-nav::-webkit-scrollbar{{display:none}}.day-nav a{{flex:0 0 46px;min-height:44px;text-align:center;text-decoration:none;color:#344054;background:#fff;border:1px solid var(--line);border-radius:11px;padding:5px 4px}}.day-nav b,.day-nav span{{display:block}}.day-nav b{{font-size:16px}}.day-nav span{{font-size:10px;color:var(--muted)}}.day-nav a.active{{color:#fff;background:var(--primary);border-color:var(--primary)}}.day-nav a.active span{{color:#dbeafe}}.filter{{flex:0 0 auto;min-height:44px;border:1px solid #f2c94c;background:#fff8db;color:#704d00;border-radius:11px;padding:10px 11px;cursor:pointer}}.filter.active{{background:#facc15}}main{{padding:16px max(14px,env(safe-area-inset-right)) calc(50px + env(safe-area-inset-bottom)) max(14px,env(safe-area-inset-left))}}.day{{margin:0 0 18px}}.day-header{{display:flex;align-items:end;justify-content:space-between;gap:12px;margin-bottom:9px;padding:0 2px}}.day-header h2{{font-size:17px;margin:2px 0 0}}.day-date{{color:var(--primary);font-weight:800;font-size:12px}}.day-summary{{text-align:right;white-space:nowrap}}.day-summary strong,.day-summary span{{display:block}}.day-summary strong{{font-size:15px}}.day-summary span{{font-size:11px;color:var(--muted);margin-top:2px}}.day.today .day-header{{border-left:4px solid var(--primary);padding-left:9px}}.timeline{{position:relative}}.timeline:before{{content:"";position:absolute;left:62px;top:18px;bottom:18px;width:2px;background:var(--line)}}.item{{--type:#64748b;position:relative;display:grid;grid-template-columns:54px 1fr;gap:18px;margin-bottom:10px}}.item:before{{content:"";position:absolute;left:58px;top:20px;width:10px;height:10px;border:3px solid var(--bg);border-radius:50%;background:var(--type);z-index:1}}.time{{padding-top:14px;font-size:11px;font-weight:700;color:#475467;text-align:right}}.item-body{{min-width:0;padding:13px 13px 11px;border-radius:15px;background:var(--card);border:1px solid #e8edf4;box-shadow:var(--shadow)}}.item.todo .item-body{{border-color:#f4cf5d}}.item-heading{{display:flex;align-items:center;gap:8px}}.item-heading h3{{font-size:16px;margin:0;min-width:0}}.type{{flex:0 0 auto;padding:3px 7px;border-radius:7px;color:var(--type);background:color-mix(in srgb,var(--type) 13%,white);font-size:11px;font-weight:800}}.note{{margin-top:8px;color:#475467;line-height:1.55;font-size:13px;white-space:pre-wrap;overflow-wrap:anywhere}}.location{{margin-top:8px;padding:8px 10px;border-radius:9px;background:#edf7ff;color:#194f86;font-size:13px;font-weight:700;overflow-wrap:anywhere}}.booking{{margin-top:8px;padding:7px 9px;border-radius:8px;background:#fff8db;color:#765600;font-size:12px}}.cost{{display:flex;align-items:center;justify-content:flex-end;gap:7px;margin-top:8px}}.cost strong{{font-size:14px}}.price-detail{{font-size:11px;color:var(--muted)}}.item-footer{{display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-top:10px}}.status{{margin-right:auto;padding:4px 8px;border-radius:99px;background:#e7edf4;color:#344054;font-size:11px;font-weight:800}}.status-待购买,.status-待预约{{background:#fef0b8;color:#704d00}}.status-已购买,.status-已预约,.status-已完成{{background:#d7f5df;color:#176b32}}.status-已取消{{background:#e5e7eb;color:#6b7280}}.actions{{display:flex;gap:5px}}.actions button,.copy-item{{display:inline-flex;align-items:center;justify-content:center;min-height:36px;border:0;background:#eef4ff;color:#2457a7;border-radius:8px;padding:7px 9px;text-decoration:none;cursor:pointer;font-size:11px}}.copy-item{{background:#f2f4f7;color:#475467}}.item.current .item-body{{border:2px solid var(--primary);box-shadow:0 10px 28px rgba(37,99,235,.16)}}.item.current:after{{content:"正在进行";position:absolute;right:10px;top:-6px;padding:3px 7px;border-radius:99px;background:var(--primary);color:#fff;font-size:10px;font-weight:800}}.item.next .item-body{{border-color:#88b5f4}}.item.next:after{{content:"下一项";position:absolute;right:10px;top:-6px;padding:3px 7px;border-radius:99px;background:#dbeafe;color:#1d4f91;font-size:10px;font-weight:800}}.map-dialog{{width:min(360px,calc(100% - 28px));border:0;border-radius:18px;padding:0;box-shadow:0 24px 70px rgba(16,24,40,.28)}}.map-dialog::backdrop{{background:rgba(15,23,42,.42);backdrop-filter:blur(2px)}}.map-panel{{padding:18px}}.map-panel h2{{margin:0;font-size:18px}}.map-place{{margin:6px 0 14px;color:var(--muted);font-size:13px;overflow-wrap:anywhere}}.map-options{{display:grid;gap:8px}}.map-options button{{min-height:48px;border:1px solid var(--line);border-radius:11px;background:#fff;color:var(--ink);font-weight:700;cursor:pointer}}.map-options button.primary{{border-color:var(--primary);background:var(--primary);color:#fff}}.map-close{{width:100%;min-height:44px;margin-top:8px;border:0;background:transparent;color:var(--muted)}}.update-toast{{position:fixed;z-index:40;left:max(14px,env(safe-area-inset-left));right:max(14px,env(safe-area-inset-right));bottom:calc(14px + env(safe-area-inset-bottom));display:flex;align-items:center;justify-content:space-between;gap:12px;max-width:520px;margin:auto;padding:12px 14px;border-radius:14px;background:#172033;color:#fff;box-shadow:0 12px 36px rgba(15,23,42,.3)}}.update-toast button{{min-height:44px;border:0;border-radius:9px;background:#fff;color:#173b73;padding:8px 14px;font-weight:800;cursor:pointer;-webkit-appearance:none}}.offline-badge{{position:fixed;z-index:30;right:max(12px,env(safe-area-inset-right));bottom:calc(12px + env(safe-area-inset-bottom));padding:6px 9px;border-radius:99px;background:#475467;color:#fff;font-size:11px;font-weight:700}}.type-交通{{--type:#2673d9}}.type-住宿{{--type:#7c4dcc}}.type-游览{{--type:#219653}}.type-餐饮{{--type:#dc7318}}.type-日常{{--type:#667085}}.type-购物{{--type:#d94f8a}}.type-其他{{--type:#80685b}}.empty{{display:none;text-align:center;color:var(--muted);padding:40px 10px}}footer{{text-align:center;color:#98a2b3;font-size:11px;padding:8px 16px 28px}}@media(max-width:420px){{h1{{font-size:23px}}.hero{{padding-bottom:18px}}.stats{{margin-top:16px}}.stat{{padding:11px 12px}}.stat strong{{font-size:18px}}.item{{grid-template-columns:50px 1fr;gap:15px}}.timeline:before{{left:58px}}.item:before{{left:54px}}.item-body{{padding:12px 12px 10px}}.actions{{gap:4px}}}}@media(max-width:370px){{.actions a{{display:none}}}}@media(display-mode:standalone){{.toolbar{{top:env(safe-area-inset-top)}}}}@media(prefers-reduced-motion:reduce){{html{{scroll-behavior:auto}}}}
</style>
</head>
<body>
<header class="hero"><div class="hero-inner"><p class="eyebrow">由 Excel 自动生成 · {esc(sheet_name)}</p><h1>{esc(title)}</h1><p class="date-range">{esc(date_text)}</p><div class="stats"><div class="stat"><span>预计总消费</span><strong>{money(total_cost)}</strong></div><div class="stat"><span>待处理事项</span><strong>{todo_count}</strong></div></div></div></header>
<nav class="toolbar"><div class="toolbar-inner"><div class="day-nav">{nav}</div><button id="todoFilter" class="filter" type="button">只看待办（{todo_count}）</button></div></nav>
<main>{''.join(sections)}<div id="empty" class="empty">没有待处理事项</div></main>
<footer>内容来源：{esc(DEFAULT_WORKBOOK)} / {esc(sheet_name)}；请修改 Excel 后重新生成。</footer>
<dialog id="mapDialog" class="map-dialog"><div class="map-panel"><h2>选择地图</h2><p id="mapPlace" class="map-place"></p><div class="map-options"><button type="button" data-map="amap">高德地图</button><button type="button" data-map="apple">Apple 地图</button><button type="button" data-map="baidu">百度地图</button></div><button id="mapClose" class="map-close" type="button">取消</button></div></dialog>
<div id="updateToast" class="update-toast" hidden><span>行程已有新版本</span><button id="reloadApp" type="button">立即刷新</button></div>
<div id="offlineBadge" class="offline-badge" hidden>当前离线</div>
<script>
(()=>{{
  if(location.search.includes('_t='))history.replaceState(null,'',location.pathname+location.hash);
  const filter=document.getElementById('todoFilter'),empty=document.getElementById('empty');
  filter.addEventListener('click',()=>{{
    const active=filter.classList.toggle('active');
    document.querySelectorAll('.item').forEach(el=>el.hidden=active&&el.dataset.todo!=='1');
    document.querySelectorAll('.day').forEach(day=>day.hidden=active&&!day.querySelector('.item[data-todo="1"]'));
    empty.style.display=active&&!document.querySelector('.item[data-todo="1"]')?'block':'none';
    filter.setAttribute('aria-pressed',String(active));
  }});
  const copyText=async button=>{{
    try{{await navigator.clipboard.writeText(button.dataset.copy)}}catch(_){{
      const area=document.createElement('textarea');area.value=button.dataset.copy;document.body.append(area);area.select();document.execCommand('copy');area.remove();
    }}
    const old=button.textContent;button.textContent='已复制';setTimeout(()=>button.textContent=old,1200);
  }};
  const mapDialog=document.getElementById('mapDialog'),mapPlace=document.getElementById('mapPlace');let selectedPlace='';
  const mapUrl=(provider,place)=>{{
    const encoded=encodeURIComponent(place);
    if(provider==='apple')return `https://maps.apple.com/?q=${{encoded}}`;
    if(provider==='baidu')return `https://api.map.baidu.com/geocoder?address=${{encoded}}&output=html&src=travel-plan`;
    return `https://uri.amap.com/search?keyword=${{encoded}}`;
  }};
  document.addEventListener('click',event=>{{
    const copy=event.target.closest('.copy,.copy-item');if(copy){{copyText(copy);return}}
    const opener=event.target.closest('.map-open');if(!opener)return;
    selectedPlace=opener.dataset.location;mapPlace.textContent=selectedPlace;
    const preferred=localStorage.getItem('preferredMap')||'amap';
    mapDialog.querySelectorAll('[data-map]').forEach(button=>button.classList.toggle('primary',button.dataset.map===preferred));
    if(mapDialog.showModal)mapDialog.showModal();else window.open(mapUrl(preferred,selectedPlace),'_blank','noopener');
  }});
  mapDialog.querySelectorAll('[data-map]').forEach(button=>button.addEventListener('click',()=>{{
    localStorage.setItem('preferredMap',button.dataset.map);window.open(mapUrl(button.dataset.map,selectedPlace),'_blank','noopener');mapDialog.close();
  }}));
  document.getElementById('mapClose').addEventListener('click',()=>mapDialog.close());
  mapDialog.addEventListener('click',event=>{{if(event.target===mapDialog)mapDialog.close()}});
  const days=[...document.querySelectorAll('.day')],links=[...document.querySelectorAll('[data-day-link]')],toolbar=document.querySelector('.toolbar'),today=new Date(),local=`${{today.getFullYear()}}-${{String(today.getMonth()+1).padStart(2,'0')}}-${{String(today.getDate()).padStart(2,'0')}}`;
  let current=days.find(day=>day.dataset.date===local)||days.find(day=>day.dataset.date>local)||days[days.length-1],scrollTarget=null,scrollTimer=0,scrollFrame=0;
  const setActive=date=>links.forEach(link=>link.classList.toggle('active',link.dataset.dayLink===date));
  if(current){{current.classList.add('today');setActive(current.dataset.date)}}
  const markCurrentItem=()=>{{
    document.querySelectorAll('.item.current,.item.next').forEach(item=>item.classList.remove('current','next'));
    const now=new Date(),items=[...document.querySelectorAll('.item[data-start]')].filter(item=>item.dataset.start&&!Number.isNaN(new Date(item.dataset.start).getTime()));
    const active=items.find(item=>{{const start=new Date(item.dataset.start),end=item.dataset.end&&item.dataset.end.slice(-5)?new Date(item.dataset.end):new Date(start.getTime()+60*60*1000);return start<=now&&now<end}});
    const next=items.find(item=>new Date(item.dataset.start)>now);
    if(active)active.classList.add('current');else if(next)next.classList.add('next');
  }};
  markCurrentItem();setInterval(markCurrentItem,60000);
  const stickyOffset=()=>toolbar.offsetHeight+(parseFloat(getComputedStyle(toolbar).top)||0)+8;
  const syncActiveDay=()=>{{
    scrollFrame=0;if(scrollTarget)return;
    const visibleDays=days.filter(day=>!day.hidden);if(!visibleDays.length)return;
    let active=visibleDays[0],line=stickyOffset();
    if(window.scrollY+window.innerHeight>=document.documentElement.scrollHeight-2)active=visibleDays[visibleDays.length-1];
    else for(const day of visibleDays){{if(day.getBoundingClientRect().top<=line)active=day;else break}}
    setActive(active.dataset.date);
  }};
  const finishNavigation=()=>{{if(!scrollTarget)return;const date=scrollTarget;scrollTarget=null;setActive(date);syncActiveDay()}};
  links.forEach(link=>link.addEventListener('click',event=>{{
    event.preventDefault();const target=document.getElementById(`day-${{link.dataset.dayLink}}`);if(!target||target.hidden)return;
    scrollTarget=link.dataset.dayLink;setActive(scrollTarget);clearTimeout(scrollTimer);
    history.replaceState(null,'',`#day-${{scrollTarget}}`);
    const top=window.scrollY+target.getBoundingClientRect().top-stickyOffset();
    window.scrollTo({{top:Math.max(0,top),behavior:'smooth'}});
    scrollTimer=setTimeout(finishNavigation,1200);
  }}));
  window.addEventListener('scroll',()=>{{
    if(scrollTarget){{clearTimeout(scrollTimer);scrollTimer=setTimeout(finishNavigation,160);return}}
    if(!scrollFrame)scrollFrame=requestAnimationFrame(syncActiveDay);
  }},{{passive:true}});
  const offlineBadge=document.getElementById('offlineBadge'),updateNetwork=()=>offlineBadge.hidden=navigator.onLine;
  window.addEventListener('online',updateNetwork);window.addEventListener('offline',updateNetwork);updateNetwork();
  if('serviceWorker' in navigator && location.protocol.startsWith('http')){{
    window.addEventListener('load',async()=>{{
      try{{
        const registration=await navigator.serviceWorker.register('./service-worker.js',{{updateViaCache:'none'}});let refreshing=false;
        const doRefresh=()=>{{if(refreshing)return;refreshing=true;location.replace(location.pathname+'?_t='+Date.now()+location.hash)}};
        const pageVersion=document.querySelector('meta[name="travel-version"]')?.content||'';
        const showUpdate=worker=>{{
          if(!worker||!navigator.serviceWorker.controller)return;
          const channel=new MessageChannel();
          channel.port1.onmessage=event=>{{
            if(event.data&&event.data.version!==pageVersion&&sessionStorage.getItem('sw-refreshed')!==event.data.version){{
              const toast=document.getElementById('updateToast');toast.hidden=false;
              document.getElementById('reloadApp').onclick=()=>{{const btn=document.getElementById('reloadApp');btn.disabled=true;btn.textContent='刷新中…';sessionStorage.setItem('sw-refreshed',event.data.version);try{{(registration.waiting||worker).postMessage({{type:'SKIP_WAITING'}})}}catch(e){{}}setTimeout(doRefresh,500);}};
            }}
          }};
          worker.postMessage({{type:'GET_VERSION'}},[channel.port2]);
        }};
        if(registration.waiting)showUpdate(registration.waiting);
        registration.addEventListener('updatefound',()=>{{const worker=registration.installing;worker.addEventListener('statechange',()=>{{if(worker.state==='installed')showUpdate(worker)}})}});
        navigator.serviceWorker.addEventListener('controllerchange',doRefresh);
        navigator.serviceWorker.addEventListener('message',event=>{{if(event.data&&event.data.type==='ACTIVATED')doRefresh()}});
        document.addEventListener('visibilitychange',()=>{{if(document.visibilityState==='visible')registration.update()}});
      }}catch(error){{console.warn('离线缓存注册失败',error)}}
    }});
  }}
}})();
</script>
</body>
</html>
'''


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip(" .")
    return cleaned or "travel"


def compatible_sheets(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        result = []
        for name in workbook.sheetnames:
            sheet = workbook[name]
            actual = [clean_text(sheet.cell(4, column).value) for column in range(1, len(HEADERS) + 1)]
            if actual == HEADERS:
                result.append(name)
        return result
    finally:
        workbook.close()


def png_chunk(chunk_type: bytes, data: bytes) -> bytes:
    return (
        struct.pack(">I", len(data))
        + chunk_type
        + data
        + struct.pack(">I", zlib.crc32(chunk_type + data) & 0xFFFFFFFF)
    )


def write_icon(path: Path, size: int) -> None:
    """生成不依赖第三方图像库的蓝色旅行图标。"""
    rows = bytearray()
    center = (size - 1) / 2
    radius = size * 0.39
    for y in range(size):
        rows.append(0)
        for x in range(size):
            distance = ((x - center) ** 2 + (y - center) ** 2) ** 0.5
            if distance <= radius:
                color = (255, 255, 255, 255)
            else:
                color = (37, 99, 235, 255)
            rows.extend(color)
    header = struct.pack(">IIBBBBB", size, size, 8, 6, 0, 0, 0)
    content = (
        b"\x89PNG\r\n\x1a\n"
        + png_chunk(b"IHDR", header)
        + png_chunk(b"IDAT", zlib.compress(bytes(rows), 9))
        + png_chunk(b"IEND", b"")
    )
    path.write_bytes(content)


def write_pwa_files(output_dir: Path, title: str, html_filename: str, content_version: str) -> None:
    icons_dir = output_dir / "icons"
    icons_dir.mkdir(parents=True, exist_ok=True)
    write_icon(icons_dir / "icon-192.png", 192)
    write_icon(icons_dir / "icon-512.png", 512)
    write_icon(icons_dir / "apple-touch-icon.png", 180)

    manifest = {
        "name": title,
        "short_name": "旅行计划",
        "description": "从 Excel 自动生成的离线旅行计划",
        "lang": "zh-CN",
        "start_url": f"./{html_filename}",
        "scope": "./",
        "display": "standalone",
        "background_color": "#f4f7fb",
        "theme_color": "#2563eb",
        "icons": [
            {"src": "icons/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "icons/icon-512.png", "sizes": "512x512", "type": "image/png"},
        ],
    }
    (output_dir / "manifest.webmanifest").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    cache_name = f"travel-{content_version}"
    assets = [
        f"./{html_filename}",
        "./manifest.webmanifest",
        "./icons/icon-192.png",
        "./icons/icon-512.png",
        "./icons/apple-touch-icon.png",
    ]
    service_worker = f'''const CACHE_NAME = {json.dumps(cache_name)};
const VERSION = {json.dumps(content_version)};
const ASSETS = {json.dumps(assets, ensure_ascii=False, indent=2)};

self.addEventListener('install', event => {{
  event.waitUntil(caches.open(CACHE_NAME).then(cache => cache.addAll(ASSETS)));
}});

self.addEventListener('message', event => {{
  if (event.data && event.data.type === 'SKIP_WAITING') self.skipWaiting();
  if (event.data && event.data.type === 'GET_VERSION' && event.ports && event.ports[0]) {{
    event.ports[0].postMessage({{version: VERSION}});
  }}
}});

self.addEventListener('activate', event => {{
  event.waitUntil(caches.keys().then(names => Promise.all(
    names.filter(name => name.startsWith('travel-') && name !== CACHE_NAME).map(name => caches.delete(name))
  )).then(() => self.clients.claim()).then(() => {{
    self.clients.matchAll().then(clients =>
      clients.forEach(client => client.postMessage({{type: 'ACTIVATED'}}))
    );
  }}));
}});

self.addEventListener('fetch', event => {{
  if (event.request.method !== 'GET') return;
  const fetchOpts = event.request.mode === 'navigate' ? {{cache: 'reload'}} : {{}};
  event.respondWith(fetch(event.request, fetchOpts).then(response => {{
    if (response.ok && new URL(event.request.url).origin === self.location.origin) {{
      const copy = response.clone();
      caches.open(CACHE_NAME).then(cache => cache.put(event.request, copy));
    }}
    return response;
  }}).catch(() => caches.match(event.request).then(cached => cached || caches.match('./{html_filename}'))));
}});
'''
    (output_dir / "service-worker.js").write_text(service_worker, encoding="utf-8")


def generate(workbook_path: Path, sheet_name: str, output_path: Path) -> None:
    title, start_day, end_day, days = read_trip(workbook_path, sheet_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    document = render_html(title, sheet_name, start_day, end_day, days)
    content_version = hashlib.sha256(document.encode("utf-8")).hexdigest()[:12]
    document = document.replace("{VERSION}", content_version)
    output_path.write_text(document, encoding="utf-8")
    write_pwa_files(output_path.parent, title, output_path.name, content_version)
    item_count = sum(len(day.items) for day in days)
    print(f"已生成 PWA：{output_path}（{len(days)} 天，{item_count} 项）")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从旅行 Excel 新版模板生成手机端 HTML")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Excel 文件路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="要生成的 Sheet 名称")
    parser.add_argument("--output", help="输出 HTML 路径；默认写入 output/index.html，并生成同目录 PWA 文件")
    parser.add_argument("--all", action="store_true", help="生成工作簿内所有新版模板 Sheet")
    parser.add_argument("--list-sheets", action="store_true", help="列出可生成的新版模板 Sheet")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).resolve()
    if not workbook_path.is_file():
        print(f"错误：找不到 Excel 文件：{workbook_path}", file=sys.stderr)
        return 1
    try:
        sheets = compatible_sheets(workbook_path)
        if args.list_sheets:
            print("\n".join(sheets) if sheets else "没有找到新版模板 Sheet")
            return 0
        if args.all:
            if args.output:
                raise ValueError("--all 不能与 --output 同时使用")
            if not sheets:
                raise ValueError("工作簿内没有新版模板 Sheet")
            for sheet_name in sheets:
                destination = Path("output") / safe_filename(sheet_name)
                generate(workbook_path, sheet_name, destination / "index.html")
        else:
            output_path = Path(args.output) if args.output else Path("output") / "index.html"
            generate(workbook_path, args.sheet, output_path)
        return 0
    except (OSError, ValueError) as error:
        print(f"错误：{error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
